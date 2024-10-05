import os
import pandas as pd
import json
from flask import Blueprint, jsonify, request, g, render_template, send_from_directory, current_app
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


report_bp = Blueprint('report', __name__)




# We need to put the warnings report code here
# Logic is to have a page that displays datatypes and allows them to select a datatype
# after selecting a datatype, they should be able to select a table that is associated with that datatype
# All this information is in the proj/config folder
#
# after selecting a table, it should display all warnings from that table
# (each table has a column called warnings, it is a varchar field and the string of warnings is formatted a certain way)
# example:
#  columnname - errormessage; columnname - errormessage2; columnname - errormessage3
# Would it have been better if we would have made it a json field? probably, but there must be some reason why we did it like this
#
# so when they select the table, we need to get all the warnings associated with that table,
#  selecting from that table where warnings IS NOT NULL
# Then we have to pick apart the warnings column text, gather all unique warnings and display to the user
# We need them to have the ability to select warnings, and then download all records from that table which have that warning

# a suggestion might be to do this how nick did the above, where he requires a query string arg,
# except we should put logic such that if no datatype arg is provided, we return a template that has links with the datatype query string arg

# example
#  <a href="/warnings-report?datatype=chemistry">Chemistry</a>
#  <a href="/warnings-report?datatype=toxicity">Toxicity</a>
# etc

@report_bp.route('/warnings-report',  methods=['GET', 'POST'])
def warnings_report():

    PROJECT_NAME = current_app.config.get('PROJECTNAME')
    DATASETS = current_app.config.get("DATASETS")

    if request.method == 'POST':
        req = request.get_json()
        datatype = req['datatype']
        table = req['table']

        if datatype is None:
            return jsonify({'no datatype'})
        elif table is None:
            table_options = DATASETS.get(datatype).get('tables')
            return jsonify({'tables': table_options})

    datatype = request.args.get('datatype')
    print(datatype)
    table = request.args.get('table')
    print(table)
    if datatype is None and table is None:
        dataset_options = DATASETS.keys()
        print("Missing fields")
        return render_template(
            'warnings-report.html',
            projectname=PROJECT_NAME,
            dataset_options=dataset_options
        )
    elif table is None:
        
        dataset_options = DATASETS.keys()
        table_options = DATASETS.get(datatype).get('tables')
        
        print("Missing fields")
        
        return render_template(
            'warnings-report.html',
            projectname=PROJECT_NAME,
            dataset_options=dataset_options,
            table_options=table_options
        )
    else:
        eng = g.eng
        sql_query = f"SELECT * FROM {table} WHERE warnings IS NOT NULL"
        tmp = pd.read_sql(sql_query, eng)
        warnings_array = tmp.warnings.apply(
            lambda x: [s.split(' - ', 1)[-1] for s in x.split(';')]).values
        unique_warnings = pd.Series(
            [item for sublist in warnings_array for item in sublist]).unique()
        df = pd.DataFrame(unique_warnings, columns=["Warnings"])
        print(df)

        warnings = df['Warnings'].tolist()

        return render_template(
            'warnings-report.html',
            projectname=PROJECT_NAME,
            datatype=datatype,
            table=table,
            warnings=warnings,
        )


@report_bp.route('/warnings-report/export/', methods=['GET', 'POST'])
def download():
    if request.method == 'POST':
        req = request.get_json()
        selected_warnings = req['warnings']
        table = req['table']

        columns = set()

        for warning in selected_warnings:
            columns_list = warning.split(' --- ')[0].split(', ')
            for column in columns_list:
                column = column.replace(' ', '')
                columns.add(column)

        print(f'columns: {columns}')


        export_name = f'{table}-export.xlsx'
        export_file = os.path.join(os.getcwd(), "export", "warnings_report", export_name)
        export_writer = pd.ExcelWriter(export_file, engine='xlsxwriter')

        warnings_data_dict = {}
        mismatched_rows = pd.DataFrame()

        for warning in selected_warnings:
            sql = f"SELECT * FROM {table} WHERE warnings LIKE %(warning)s"
            data = pd.read_sql(sql, g.eng, params={'warning': f"%{warning}%"})
            
            split_warnings = data['warnings'].str.split(';').apply(pd.Series, 1).stack()
            split_warnings.index = split_warnings.index.droplevel(-1)
            split_warnings.name = 'warning'

            del data['warnings']
            data = data.join(split_warnings)

            matching_rows = data[data['warning'] == warning]
            mismatched_rows = pd.concat([mismatched_rows, data[~data.index.isin(matching_rows.index)]])
            warnings_data_dict[warning] = matching_rows

        for index, row in mismatched_rows.iterrows():
            warning = row['warning']
            if warning in warnings_data_dict:
                warnings_data_dict[warning] = pd.concat([warnings_data_dict[warning], pd.DataFrame([row])])

        yellow_fill = PatternFill(
            start_color='00FFFF00',
            end_color='00FFFF00',
            fill_type='solid'
        )

        with export_writer:
            for index, warning in enumerate(warnings_data_dict):
                sheetname = f"error{index+1}"
                warnings_data_dict[warning].to_excel(export_writer, sheet_name = sheetname, index = False)


        for index, warning in enumerate(warnings_data_dict):
            
            sheetname = f"error{index+1}"
            
            wb = load_workbook(export_file)
            
            columns_list = warning.split(' --- ')[0].split(',')
            
            for column in columns_list:
                column = column.replace("'", '')
                column = column.replace(' ', '').lower()
            
                target_column = None
                target_header = column

                for col_idx, col in enumerate(wb[sheetname].iter_cols(max_row=1)):
                    if col[0].value == target_header:
                        target_column = col_idx + 1
                        break
                
                if target_column is not None:
                    for row in wb[sheetname].iter_rows(min_col=target_column, max_col=target_column):
                        for cell in row:
                            cell.fill = yellow_fill
                            warning_msg = warning.split(' --- ')[1]
                            cell.comment = Comment(warning_msg, "Checker")
                    wb.save(export_file)

        return send_from_directory(os.path.join(os.getcwd(), "export", "warnings_report"), export_name, as_attachment=True)
    


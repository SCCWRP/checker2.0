import os, shutil
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Border, Side, PatternFill

from flask import session
from math import floor


def mark_workbook(all_dfs, excel_path, errs, warnings):
    assert session.get('submission_dir') is not None, "function - mark_workbook - session submission dir is not defined."
    orig_filename = excel_path.rsplit('/', 1)[-1]
    filename = orig_filename.rsplit('.',1)[0]
    ext = orig_filename.rsplit('.',1)[-1]
    marked_path = os.path.join(session.get('submission_dir'), f"{filename}-marked.{ext}")
    
    # copy the excel file and have "marked" in the name, and we will mark this excel file
    shutil.copy(excel_path, marked_path)

    # No empty errors allowed otherwise it crashes
    errs = [e for e in errs if len(e) > 0]

    errs_cells = dict()
    for table in set([e.get('table') for e in errs]):
        errs_cells[table] = []
        [    
            errs_cells[table].append(
                {
                    'row_index': r,
                    'column_index': all_dfs[table].columns.get_loc(str(col).strip().lower()),
                    'message': e.get('error_message')
                }
            )
            for e in errs
            for col in e.get('columns').split(',')
            for r in e.get('rows')
            if e.get('table') == table
        ]

    warnings_cells = dict()
    for table in set([w.get('table') for w in warnings]):
        warnings_cells[table] = []
        [    
            warnings_cells[table].append(
                {
                    'row_index': r,
                    'column_index': all_dfs[table].columns.get_loc(str(col).strip().lower()),
                    'message': f"{w.get('error_message')} (Warning)"
                }
            )
            for w in warnings
            for col in w.get('columns').split(',')
            for r in w.get('rows')
            if w.get('table') == table
        ]

    
    # for errors
    redFill = PatternFill(
        start_color='FF8585',
        end_color='FF8585',
        fill_type='solid'
    )

    # for warnings
    yellowFill = PatternFill(
        start_color='00FFFF00',
        end_color='00FFFF00',
        fill_type='solid'
    )


    wb = load_workbook(marked_path)

    for sheet in wb.sheetnames:
        # Mark warnings first - this way if there are an error and warning in the same cell, the error will be shown
        for coord in warnings_cells.get(sheet) if warnings_cells.get(sheet) is not None else []:
            colindex = coord.get('column_index')
            wb[sheet][f"{chr(65 +  (floor(colindex/26) - 1)  ) if colindex >= 26 else ''}{chr(65 + (colindex % 26))}{int(coord.get('row_index'))}"].fill = yellowFill 
            wb[sheet][f"{chr(65 +  (floor(colindex/26) - 1)  ) if colindex >= 26 else ''}{chr(65 + (colindex % 26))}{int(coord.get('row_index'))}"].comment = Comment(coord.get('message'), "Checker")
        
        for coord in errs_cells.get(sheet) if errs_cells.get(sheet) is not None else []:
            
            # the workbook sheet or whatever its called accesses the cells of the excel file not with the numeric indexing like pandas but rather that letter indexing thing
            # like "Cell A1" and stuff like that

            # So the gigantic disgusting f string f"{chr(65 +  (math.floor(colindex/26) - 1)  ) if colindex >= 26 else ''}{chr(65 + (colindex % 26))}{coord.get('row_index')}"
            # is to convert from pandas indexing to the letter indexing style thing

            colindex = coord.get('column_index')
            wb[sheet][f"{chr(65 +  (floor(colindex/26) - 1)  ) if colindex >= 26 else ''}{chr(65 + (colindex % 26))}{int(coord.get('row_index'))}"].fill = redFill 
            wb[sheet][f"{chr(65 +  (floor(colindex/26) - 1)  ) if colindex >= 26 else ''}{chr(65 + (colindex % 26))}{int(coord.get('row_index'))}"].comment = Comment(coord.get('message'), "Checker")
        
    wb.save(marked_path)

    return marked_path



def format_existing_excel(file_path_or_bytes_object, header_row = 1, cushion = 5, freeze_headers = True):
    
    assert isinstance(file_path_or_bytes_object, (str, BytesIO)), "file_path_or_bytes_object must be a string or BytesIO"
    # Load the workbook and iterate through sheets
    
    if isinstance(file_path_or_bytes_object, BytesIO):
        file_path_or_bytes_object.seek(0)
        workbook = load_workbook(filename=BytesIO(file_path_or_bytes_object.read()))
    else:
        workbook = load_workbook(file_path_or_bytes_object)

    # Define a light grey fill
    grey_fill = PatternFill(
        start_color='00BABABA', 
        end_color='00BABABA', 
        fill_type='solid'
    )
    # Define a light fill for zebra striping
    stripe_fill = PatternFill(
        start_color='00DBDBDB', 
        end_color='00DBDBDB', 
        fill_type='solid'
    )
    # Define a border for the table body
    table_body_border = Border(
        left=Side(border_style='thin', color='00AAAAAA'),
        right=Side(border_style='thin', color='00AAAAAA'),
        top=Side(border_style='thin', color='00AAAAAA'),
        bottom=Side(border_style='thin', color='00AAAAAA')
    )
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Apply zebra striping starting from the row after the header
        for i, row in enumerate(sheet.iter_rows(min_row=header_row+1), start=header_row+1):
            for cell in row:
                cell.border = table_body_border
                if i % 2 == 0:  # For even row numbers
                    cell.fill = stripe_fill

        if freeze_headers == True:
            # Freeze the row just below the header row
            freeze_cell = 'A' + str(header_row + 1)
            sheet.freeze_panes = freeze_cell

        # Apply formatting to the specified header row
        for cell in sheet[header_row]:  # Use the header_row parameter
            cell.font = Font(bold=True)
            cell.border = Border(
                top=Side(style='thin'), 
                bottom=Side(style='thin'),
                left=Side(style='thin'), 
                right=Side(style='thin')
            )
            cell.fill = grey_fill

        # Set the column widths based on max length in column
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[header_row-1:])
            adjusted_width = max_length + cushion  # Add cushion for a little extra space
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Apply filters to the specified header row
        if sheet.max_row >= header_row:  # Check if the header_row is within the data range
            sheet.auto_filter.ref = f"{sheet.dimensions.split(':')[0]}:{sheet.dimensions.split(':')[1]}"

    # Save the workbook
    if isinstance(file_path_or_bytes_object, BytesIO):
        outstream = BytesIO()
        workbook.save(outstream)
        outstream.seek(0)
        return outstream
    else:
        workbook.save(file_path_or_bytes_object)
        return
        
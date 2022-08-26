from pandas.io import excel
from flask import render_template, request, jsonify, current_app, Blueprint, session, g, send_file
from werkzeug.utils import secure_filename
from sqlalchemy import create_engine, text, exc, Table, MetaData
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from pandas import DataFrame
import re
import os
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows

# dynamic lookup lists to template

templater = Blueprint('templater', __name__)
@templater.route('/templater', methods = ['GET', 'POST']) # this will be added to the index.html file to dynamically call the lookup lists to each template
# consider using the app.datasets dictionary to generalize the code better
def template():
    system_fields = current_app.system_fields
    datatype = request.args.get("datatype")

    # we are assuming that the datatype arg is one of the keys on the dtypes dictionary
    # an assert statement might be good
    tables = current_app.dtypes.get('datatype').get('tables')

    eng = g.eng

    for tablename in tables:
        lookup_sql = f"""
            SELECT
                kcu.column_name, 
                ccu.table_name AS foreign_table_name,
                ccu.column_name AS foreign_column_name 
            FROM 
                information_schema.table_constraints AS tc 
                JOIN information_schema.key_column_usage AS kcu
                ON tc.constraint_name = kcu.constraint_name
                AND tc.table_schema = kcu.table_schema
                JOIN information_schema.constraint_column_usage AS ccu
                ON ccu.constraint_name = tc.constraint_name
                AND ccu.table_schema = tc.table_schema
            WHERE tc.constraint_type = 'FOREIGN KEY' 
            AND tc.table_name='{tablename}'
            AND ccu.table_name LIKE 'lu_%%';
        """

        # fkeys = foreign keys
        fkeys = pd.read_sql(lookup_sql, eng)
        # dont check lookup lists for columns that are not in unified table
        fkeys = fkeys[fkeys.column_name.isin(dataframe.columns)]
        



    df = df[df['table_from'].str[:4].str.contains('tbl_')]
    grouped_df = df.groupby('table_from')

    for key, item in grouped_df:
        print(grouped_df.get_group(key), "\n\n")

    descr_grouped_df = grouped_df.describe()
    print(descr_grouped_df)
    
    # dtypes is a list of datatypes for empa (in order by SOP No.)
    dtype = [
        'logger',
        'discretewq',
        'nutrients_lab',
        'nutrients_field',
        'edna_field',
        'edna_lab',
        'sedimentgrainsize_field',
        'sedimentgrainsize_lab',
        'benthicinfauna_field',
        'benthicinfauna_lab',
        'macroalgae',
        'sav',
        'bruv_field',
        'bruv_lab',
        'fishseines',
        'crabtrap',
        'vegetation',
        'feldspar'
    ]

    # if/else statements to create list of tables contained within template for each datatype
    # SOP 1 WQ Logger
    if datatype == 'logger':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_wq_logger_metadata',
            'tbl_logger_ctd_data',
            'tbl_logger_mdot_data',
            'tbl_logger_troll_data',
            'tbl_logger_tidbit_data'
            ]
        file_prefix = 'SOP_1_WQ_LOGGER'
    # SOP 2 Discrete WQ
    elif datatype == 'discretewq':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_waterquality_metadata',
            'tbl_waterquality_data'
            ]
        file_prefix = 'SOP_2_DISCRETE_WQ'
    # SOP 3 Nutrients
    elif datatype == 'nutrients_field':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_nutrients_metadata'
            ]
        file_prefix = 'SOP_3_NUTRIENTS_FIELD'
    elif datatype == 'nutrients_lab':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_nutrients_labbatch_data',
            'tbl_nutrients_data'
            ]
        file_prefix = 'SOP_3_NUTRIENTS_LAB'
    # SOP 4 eDNA
    elif datatype == 'edna_field':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_edna_metadata'
            ]
        file_prefix = 'SOP_4_EDNA_FIELD'
    elif datatype == 'edna_lab':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_edna_water_labbatch_data',
            'tbl_edna_sed_labbatch_data',
            'tbl_edna_data'
            ]
        file_prefix = 'SOP_4_EDNA_LAB'
    # SOP 5 Sediment Grain Size
    elif datatype == 'sedimentgrainsize_field':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_sedgrainsize_metadata'
            ]
        file_prefix = 'SOP_5_SEDIMENTGRAINSIZE_FIELD'
    elif datatype == 'sedimentgrainsize_lab':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_sedgrainsize_data',
            'tbl_sedgrainsize_labbatch_data'
            ]
        file_prefix = 'SOP_5_SEDIMENTGRAINSIZE_LAB'
    # SOP 6 Benthic Infauna
    elif datatype == 'benthicinfauna_field':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_benthicinfauna_metadata'
            ]
        file_prefix = 'SOP_6_BENTHICINFAUNA_FIELD'
    elif datatype == 'benthicinfauna_lab':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_benthicinfauna_labbatch',
            'tbl_benthicinfauna_abundance',
            'tbl_benthicinfauna_biomass'
            ]
        file_prefix = 'SOP_6_BENTHICINFAUNA_LAB'
    # SOP 7 Macroalgae
    elif datatype == 'macroalgae':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_macroalgae_sample_metadata',
            'tbl_algaecover_data',
            'tbl_floating_data'
            ]
        file_prefix = 'SOP_7_MACROALGAE'
    # SOP 7 SAV
    elif datatype == 'sav':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_sav_metadata',
            'tbl_savpercentcover_data'
            ]
        file_prefix = 'SOP_7_SAV'
    # SOP 8 BRUV
    elif datatype == 'bruv_field':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_bruv_metadata',
            ]
        file_prefix = 'SOP_8_BRUV_FIELD'
    elif datatype == 'bruv_lab':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_bruv_data'
            ]
        file_prefix = 'SOP_8_BRUV_LAB'
    # SOP 9 Fish Seines
    elif datatype == 'fishseines':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_fish_sample_metadata',
            'tbl_fish_abundance_data',
            'tbl_fish_length_data'
            ]
        file_prefix = 'SOP_9_FISHSEINES'
    # SOP 10 Crab Trap
    elif datatype == 'crabtrap':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_crabtrap_metadata',
            'tbl_crabfishinvert_abundance',
            'tbl_crabbiomass_length'
            ]
        file_prefix = 'SOP_10_CRABTRAP'
    # SOP 11 Vegetation
    elif datatype == 'vegetation':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_vegetation_sample_metadata',
            'tbl_vegetattivecover_data',
            'tbl_epifauna_data'
            ]
        file_prefix = 'SOP_11_VEGETATION'
    # SOP 13 Feldspar
    elif datatype == 'feldspar':
        tbls = [
            'tbl_protocol_metadata',
            'tbl_feldspar_metadata',
            'tbl_feldspar_data'
            ]
        file_prefix = 'SOP_13_FELDSPAR'
    else:
        tbls = []

    #initialize list for primary and foreign keys extracted from database
    keys_from_db = []
    
    for key, item in grouped_df:
        if key in tbls:
            print(key)
            print(item['pg_get_constraintdef'])
            keys_from_db.append(item['pg_get_constraintdef'])

    list_of_keys = []
    list_of_lu_needed = []

    for fkey in keys_from_db:
        print(f"fkey ----- {fkey}")
        
        for lu in fkey: 
            print(f"lu --- {lu}")
            print("appending list of keys: ")
            list_of_keys.append(lu)
    print("LIST OF KEY")
    print(list_of_keys)

    primarykeylist = list()
    tmplist = list()

    for line in list_of_keys:
        #GET PRIMARY KEY COLUMNS FROM THE DATABASE AND ADD TO PRIMARYKEYCOLUMN LIST
        if "PRIMARY" in line: 
            print(line)
            tmpprimary = line.split()[2:]
            for primary in tmpprimary:
                primary = re.sub('\(|\)|\,|', '', primary)
                primarykeylist.append(primary)
        for element in line.split():
            if element.startswith('lu'):
                ## ignoring this part because it seems like it relates to smc, dont worry about it tbh
                print('The following is a lookup list: ', element)
                ## ignoring this part because it seems like it relates to smc, dont worry about it tbh
                #if element.split('(')[0] in ['lu_station', 'lu_stations', 'lu_organismalgae']:
                    #continue
                list_of_lu_needed.append(element.split('(')[0])
            if element.startswith('('):
                foreignkey = element.strip('()')
                #GET FOREIGN KEY COLUMNS FROM THE DATABASE AND ADD TO TMPLIST
                tmplist.append(foreignkey)

    print("THIS IS FOREIGN KEY COLUMNS(TBLS) + PRIMARY KEY COLUMNS(LOOK_UP) LIST ")
    print(tmplist)
    #Currently at this point, the tmplist has only foreign key columns for the tabs/tbls. This list will be updated with the primary key columns(look-up list) later.
    print("THIS IS PRIMARY KEY COLUMN(TBLS) LIST")
    print(primarykeylist)
    list_lu_needed = list(set(list_of_lu_needed))
    print(list_of_lu_needed)

    lu_list_df = []

    for i, lu_list in enumerate(list_of_lu_needed):
        print(i, lu_list)
        sql = eng.execute("SELECT * FROM " + lu_list)
        sql_df = DataFrame(sql.fetchall())
        sql_df.columns = sql.keys()
        lu_list_df.append(sql_df)
        del sql_df

    del i, lu_list

    # print("export file")
    # print(f"datatype: {datatype}, {datatype.upper()}")
    # excel_file = f"{os.getcwd()}/export/routine/{file_prefix}-TEMPLATE.xlsx"
    # excel_writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    # print("creating workbook")
    # workbook = excel_writer.book
    # print(" ===== created workbook ===== ")
    # print("reading the template in")

    # Bug: When the application crashes, open files that should be appended seem to be corrupted since the file was not closed.
        ## Not a problem for now since I changed the flow of the code.
    # Bug Fix Attempt:  Include try... finally approach to explicitly close files -- -include this later
    print("===== reading in template file =====")
    print(f"datatype: {datatype}")
    print(datatype in dtype)
    if datatype in dtype:
        print(f"get template file for datatype: {datatype}")
        e_file = f"{os.getcwd()}/export/routine/{file_prefix}-TEMPLATE.xlsx"
        print(f"this is the file: {e_file}")
        ## The following returns an OrderedDict
        xls = pd.read_excel(e_file, sheet_name=None, engine='openpyxl')
        print("ordered dictionary keys: ")
        print(xls.keys())
        print(" made it here!!!  ")
    else: 
        print('invalid datatype')
    print('xls template for data type has been read')

    print("The following are the sheet_names of xls: ")
    print(xls.keys())

    #initialize lists
    list_of_df = []
    original_df = []

    #enumerate the function to give list of the sheet_names - using OrderedDict keys()
    for i, sheet in enumerate(xls.keys()):
        print("list_of_df: ")
        print(list_of_df)
        #list_of_df.append(pd.read_excel(xls, sheet_name = sheet))
        list_of_df.append(pd.read_excel(e_file, sheet_name = sheet))
        print("Printing sheet name: ")
        print(sheet)
        print("Printing sheet contents: ")
        #print(pd.read_excel(xls, sheet_name = sheet))
        print(pd.read_excel(e_file, sheet_name = sheet))
        original_df.append(pd.read_excel(e_file, sheet_name = sheet))
        #lowercase
        #list_of_df[i].columns = map(str.lower(), list_of_df[i].columns)
        list_of_df[i].columns = [str(x).lower() for x in list_of_df[i].columns]

    del i, sheet
    print("removed intermediate variables...")

    ## GRAB LOOKUP LISTS RFOM DB
    print("These are original sheets :")
    print(list_of_df)
    lookup_list = pd.read_sql("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME LIKE 'lu%%'",eng)
    lookup_list = lookup_list['table_name'].tolist()
    
    ## PRIMARY KEY 
    primarykey = dict()
    # x : lookuplist variable for findprimary()
    def findprimary(x):
        for i in lookup_list:
            # meta is an initialized variable for MetaData()
            s = Table(i,meta,autoload =True, autoload_with = eng)
            lpkeys = list()
            for pk in s.primary_key:
                pri = pk.name
                lpkeys.append(pri)
                primarykey[i] = lpkeys
        pkeytable = pd.DataFrame.from_dict(primarykey,orient = 'index')
        pkeytable.reset_index(level = 0, inplace = True)
        for i in pkeytable.columns:
            if i =='index':
                pkeytable.rename(columns = {i :"table"},inplace = True)
            else:
                pkeytable.rename(columns = {i: "primarycolumn_" + str(i)},inplace = True)
        pkeytable = pkeytable.sort_values('table')
        return pkeytable

    pkeytable = findprimary(lookup_list)
    print("-----------------------------")
    print(" Printing pkey table below...")
    print("-----------------------------")
    print(pkeytable)

    def highlight_primary_key_column(dict):
        columns_to_highlight = list()
        for i in dict.keys():
            for index,row in pkeytable.iterrows():
                if i == row['table']:
                    df = dict[i]
                    for ii in df.columns:
                        if ii == row['primarycolumn_0']: 
                            columns_to_highlight.append(ii)
        print(columns_to_highlight)
        return columns_to_highlight

    ## RETRIEVE LOOKUP LISTS FROM THE DATABASE
    ###################
    columnlist = list()
    columndict = dict()
    for i, lu_list in enumerate(list_of_lu_needed):
        # It's easier to use pd.read_sql in my opinion
        sql_df = pd.read_sql("SELECT * FROM " + lu_list, eng)
        print("retrieved sql_df for lookups...")
        #sql_df = sql_df[[x for x in sql_df.columns if x not in app.system_fields]] # no app.system_fields here
        sql_df = sql_df[[x for x in sql_df.columns if x not in system_fields]]
        templater = dict({lu_list:sql_df})
        columns_to_highlight = highlight_primary_key_column(templater)
        #columnlist is the primary keys for the lu lists
        columnlist.extend(columns_to_highlight)
        for value in columns_to_highlight:
            test = dict()
            test[columns_to_highlight[0]] = sql_df[value]
            columndict.update(test)
    columnlist.extend(tmplist)
    #export original_df to the excel file
    print("exporting to original_df to excel file")
    # create excel_writer object here?
    print("==================================================")
    print("==================================================")
    print("==================================================")
    #print("export file")
    excel_file = f"{os.getcwd()}/export/routine/{file_prefix}-TEMPLATE.xlsx"
    excel_writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    #print("creating workbook object")
    #workbook = excel_writer.book
    #print(" ===== created workbook ===== ")
    
    print("this is what original_df is populated with: ")
    print(original_df)

    # e_file : file path used to read and write to workbook for any updates
    for i, sheet in enumerate(xls.keys()):
        print(f"length of original_df: {len(original_df)}")
        print(f"i: {i}")
        print(f"sheet: {sheet}")
        print(f"e_file: {e_file}")
        original_df[i].to_excel(excel_writer, sheet_name = sheet, startrow = 1,index = False,header = False)
        print(" please print so i know it did in fact work as expected ")
        print("========================================================")
        workbook = excel_writer.book
        worksheet = excel_writer.sheets[sheet]
        print("fixed yet another baby bug")
        #bold indicated foreign keys, otherwise not bold 
        #format 1 is for FOREIGN KEY COLUMNS
        format1= workbook.add_format({'bold':False, 'text_wrap': True,'fg_color': '#D7D6D6'}) 
        format1.set_align('center')
        format1.set_align('vcenter')
        format1.set_rotation(90)
        #format 2 is for REGULAR COLUMNS
        format2= workbook.add_format({'bold':False, 'text_wrap': True})
        format2.set_align('center')
        format2.set_align('vcenter')
        format2.set_rotation(90)
        #format 3 is for PRIMARY KEY COLUMNS
        format3= workbook.add_format({'bold':True, 'text_wrap': True})
        format3.set_align('center')
        format3.set_align('vcenter')
        format3.set_rotation(90)
        #format 4 is for PRIMARY KEY AND ALSO FOREIGN KEY COLUMNS
        format4= workbook.add_format({'bold':True, 'text_wrap': True,'fg_color': '#D7D6D6'})
        format4.set_align('center')
        format4.set_align('vcenter')
        format4.set_rotation(90)
        if sheet == 'Instructions':
            print(f"sheet: {sheet}")
            continue
        for col_num,col_name in enumerate(original_df[i].columns.values):
            if   (col_name.lower() in primarykeylist) & (col_name.lower() not in columnlist)  : 
                worksheet.write(0,col_num,col_name,format3)
                worksheet.set_row(0,170)
            elif (col_name.lower() in primarykeylist) & (col_name.lower() in columnlist):
                worksheet.write(0,col_num,col_name,format4)
                worksheet.set_row(0,170)
            elif (col_name.lower() in columnlist) & (col_name.lower() not in primarykeylist):
                worksheet.write(0,col_num,col_name,format1)
                worksheet.set_row(0,170)
            else:
                worksheet.write(0,col_num,col_name,format2)
                worksheet.set_row(0,170)
    del i, sheet

    ### ADD THE LOOKUPS TO THE FILE: 
    #Add the look_up lists to the excel file
    for i, lu_list in enumerate(list_of_lu_needed):
        sql_df = pd.read_sql("SELECT * FROM " + lu_list, eng)
        #sql_df = sql_df[[x for x in sql_df.columns if x not in app.system_fields]] # NO app.system_fields
        sql_df = sql_df[[x for x in sql_df.columns if x not in system_fields]]
        sql_df.to_excel(excel_writer, sheet_name = lu_list,startrow =1,index =False,header = False)
        workbook = excel_writer.book
        worksheet = excel_writer.sheets[lu_list]
        #change set tab color to hex #BFBFBF (dark grey) for lu lists
        worksheet.set_tab_color('#A6A6A5')
        format1= workbook.add_format({'bold':True, 'text_wrap': True,'fg_color': '#D7D6D6'})
        format1.set_align('center')
        format1.set_align('vcenter')
        format1.set_rotation(90)
        format2= workbook.add_format({'bold':False, 'text_wrap': True,'valign':'center'})
        format2.set_rotation(90)
        format2.set_align('center')
        format2.set_align('vcenter')
        for col_num,col_name in enumerate(sql_df.columns.values):
            if col_name in columnlist:
                worksheet.write(0,col_num ,col_name,format1)
                worksheet.set_row(0,120)
            else:
                worksheet.write(0,col_num ,col_name,format2)
                worksheet.set_row(0,120)
        print("saving excel writer") 
    print(" ===== saved ===== ")
    excel_writer.save()

    ###############################################################################
    #               D A T A   V A L I D A T I O N  D E S C R I P T I O N          #
    ###############################################################################
    print("Begin Data Validation Descriptions...")
    # populate df_glossary from database
    sql = eng.execute("SELECT * FROM tbl_glossary;")
    sql_df = DataFrame(sql.fetchall())
    sql_df.columns = sql.keys()
    df_glossary = sql_df
    del sql_df

    # SUBSET GLOSSARY TO SPECIFIC DATATYPE
    df = df_glossary[df_glossary['datatype'] == datatype]
    #file_prefix= df.template_prefix.unique()[0].upper() # this should already be filled from earlier in the code 

    # GROUP BY SHEET, THE DATAFRAME IS ALREADY FILTERED TO DATATYPE RELEVANT VALUES ONLY
    grouped_df = df.groupby(['sheet'])
    gb_df = grouped_df.groups
    key_df = gb_df.keys()

    # LOADING WORKBOOK TO APPEND
    wb = openpyxl.load_workbook(e_file)

    for key, values in gb_df.items():
        print(key)
        print("\n")
        print(values)
        print("\n")
        sh = wb[key]
        print(sh)
        print("\n")
        if key == 'glossary':
            print('----- updating the glossary table -----')
            glossary_contents = df[['sheet','field_name','definition','field_type','format_text','unit']]
            sh.delete_rows(2, sh.max_row-1)
            for r in dataframe_to_rows(glossary_contents, index=False, header=False):
                sh.append(r)
        tmp = df[df['sheet'] == key]
        print(tmp)
        print("\n")
        field_df_dict = dict(zip(tmp['field_name'], tmp['definition']))
        #print(field_df_dict)
        print("\n")
        n = len(tmp.field_name.tolist()) - 1
        print(f"n: {n}")
        for row in sh.iter_rows(min_row=1, min_col=1, max_row=1, max_col=n):
            for cell in row:
                print(cell)
                print(cell.value, end=" ")
                dv = DataValidation()
                s = str(field_df_dict[cell.value]) if cell.value is not None else None
                #maybe just dont print it
                #print(f"s: {s}")
                dv.prompt = s
                sh.add_data_validation(dv)
                dv.add(cell)
                print("data validation description added to cell")

    wb.save(e_file)
    wb.close()

    print(" ====== saved =====")
    # probably better to generalize with the below code, but for testing right now let's just call the exact file path and modify
    #return send_file(f'{os.getcwd()}/export/{datatype.upper()}-TEMPLATE.xlsx', as_attachment=True, attachment_filename=f'{datatype.upper()}-TEMPLATE.xlsx')
    #return send_file(f'{os.getcwd()}/export/routine/{datatype.upper()}-TEMPLATE.xlsx', as_attachment=True, attachment_filename=f'{datatype.upper()}-TEMPLATE.xlsx')
    return send_file(f"{os.getcwd()}/export/routine/{file_prefix}-TEMPLATE.xlsx", as_attachment=True, attachment_filename=f'{file_prefix}-TEMPLATE.xlsx')
    





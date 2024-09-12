# Templater routine dynamically calls lookup lists from database to populate into the data templates.
# The updated templates will be outputted to the user with all tables and the lookup lists. 
#below is from microplastics
#from flask import send_from_directory, render_template, request, redirect, Response, jsonify, send_file, json, current_app
#below is from empa main.py
from flask import request, current_app, Blueprint, g, send_file, make_response, flash, render_template
from sqlalchemy import Table, MetaData
import pandas as pd
from pandas import DataFrame
import re
import os
import openpyxl
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.styles import PatternFill, Font, Border, Side, Color
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment
from io import BytesIO

from .utils.db import primary_key, foreign_key_detail, get_column_comments

# dynamic lookup lists to template
# skip the formatting

templater = Blueprint('templater', __name__)


@templater.route('/templates', methods = ['GET', 'POST']) # this will be added to the index.html file to dynamically call the lookup lists to each template
@templater.route('/templater', methods = ['GET', 'POST']) # this will be added to the index.html file to dynamically call the lookup lists to each template
# consider using the app.datasets dictionary to generalize the code better
def template():
    system_fields = current_app.system_fields
    datatype = request.args.get("datatype")

    if datatype not in current_app.datasets.keys():
        if datatype is not None:
            flash(f"Datatype {datatype} not found")
        return render_template(
            "templates.jinja2",
            datasets = current_app.datasets,
            project_name = current_app.project_name,
            background_image = current_app.config.get("BACKGROUND_IMAGE")
        )
    
    # All tables in the dataset
    tbls = current_app.datasets.get(datatype)['tables']
    
    file_prefix = datatype.upper()
    print(current_app.datasets.keys())

    static_template = current_app.datasets.get(datatype).get('template_filename')
    if static_template is not None:
        print("Static template requested")
        return send_file(
            os.path.join(os.getcwd(), "export", "data_templates", static_template), as_attachment=True, download_name=f'{static_template}'
        )
    
    eng = g.eng

    tabs_dict = {}

    for tbl in tbls:
        print('tbl')
        print(tbl)
        # foreign key detail returns a dictionary in records fashion rather than just columns and which tables they reference
        # foreign key detail gives the name of the referenced column as well
        
        pkey_fields = primary_key(tbl, eng)
        fkey_detail = foreign_key_detail(tbl, eng)
        print('fkey_detail')
        print(fkey_detail)
        tabs_dict[tbl] = {
            'pkey_fields' : pkey_fields,
            'foreign_key_detail' : fkey_detail,
            'foreign_key_tables' : list({info['referenced_table'] for info in fkey_detail.get(tbl, {}).values()}),
            'constrained_columns' : list(fkey_detail.get(tbl, {}).keys())
        }
        
        print("tabs dict")
        print( tabs_dict[tbl] )
    
    lookup_tables = [t for v in tabs_dict.values() for t in v.get('foreign_key_tables')]
    
    # Build the dictionary which will be the main object that creates the data template
    xls = {
        **{
            'Instructions': pd.DataFrame(
                
                    {
                        'How to use:': [
                            "Information about this spreadsheet:",
                            "SCCWRP spreadsheets follow a standard format, each consisting of several sheets: data templates, lookup lists, and a glossary.",
                            "Metadata for each column can be found by selecting the header cell for that column, or in the glossary sheet. Please do not add or rename columns. Use note columns to provide additional information or context.",
                            "Questions or comments? Please contact Paul Smith at pauls@sccwrp.org"
                        ]
                    }
                
            )
        },
        **{
            table: pd.DataFrame(
                columns=
                [
                    *[
                        x for x in pd.read_sql(
                            """
                                SELECT {} FROM {} LIMIT 1
                            """.format(
                                ','.join(pd.read_sql(f"SELECT column_name FROM column_order WHERE table_name = '{table}' ORDER BY custom_column_position;", eng).column_name.tolist()),
                                table
                            ),
                            eng
                        ).columns.to_list()
                        if x not in system_fields
                    ]
                ]
            ) for table in tbls # remember, tbls was defined towards the beginning, and represents all tables in the dataset
        },
        **{
            'glossary': pd.read_sql(f"""SELECT * FROM vw_template_glossary WHERE tablename IN ('{"','".join(tbls)}') ;""", eng)
        },
        **{
            lu_name: pd.read_sql(f"SELECT * from {lu_name}", eng).drop(columns=['objectid'], errors = 'ignore')
            for lu_name in list(set(lookup_tables))
        }
    }
    
    excel_blob = BytesIO()
        
    with pd.ExcelWriter(excel_blob, engine='openpyxl') as writer:
        
        # Set the correct font size for the column headers, according to the config file
        try:
            COLUMN_HEADER_FONT_SIZE = current_app.config.get("TEMPLATE_COLUMN_HEADER_FONT_SIZE", 12)
            COLUMN_COMMENT_FONT_SIZE = current_app.config.get("TEMPLATE_COLUMN_COMMENT_FONT_SIZE", 12)
        except Exception as e:
            
            print(
                "Warning: Could not set custom header font size for data submission template - most likely there is an error in the app configuration (wrong datatype for TEMPLATE_COLUMN_HEADER_FONT_SIZE?)"
            )
            
            COLUMN_HEADER_FONT_SIZE = 12
            COLUMN_COMMENT_FONT_SIZE = 12
        
        # Get the custom column header fill
        CUSTOM_COLUMN_HEADER_FILL = current_app.config.get("TEMPLATE_COLUMN_HEADER_FILL", None)
        CUSTOM_COLUMN_HEADER_BORDER_COLOR = current_app.config.get("TEMPLATE_COLUMN_HEADER_BORDER_COLOR", "#000000")
        
        CUSTOM_COLUMN_COMMENT_FILL = current_app.config.get("TEMPLATE_COLUMN_COMMENT_FILL", None)
        CUSTOM_COLUMN_COMMENT_BORDER_COLOR = current_app.config.get("TEMPLATE_COLUMN_COMMENT_BORDER_COLOR", "#000000")
        
        # Gray highlight format
        FKEY_HIGHLIGHT = PatternFill(start_color="D7D6D6", end_color="D7D6D6", fill_type="solid")
        PKEY_BOLD_FONT = Font(bold=True,size=COLUMN_HEADER_FONT_SIZE)
        
        try:
            rotation = int(current_app.config.get("TEMPLATE_COLUMN_HEADER_ROTATION", 90))
            COL_HEADER_ROTATION = Alignment(text_rotation = rotation , horizontal='center', vertical='center')
        except Exception as e:
            print("Warning: Couldnt set custom column header rotation - likely an error in the app configuration - defaulting to 90")
            COL_HEADER_ROTATION = Alignment(text_rotation=90, horizontal='center', vertical='center')
            
        
        # Define a light red fill
        DATA_VALIDATION_ERROR_FILL = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

        # Fetch column comments if config option is True
        INCLUDE_COMMENTS = current_app.config.get("INCLUDE_COLUMN_COMMENTS", False)
        
        # Start the offset at 0, add one 
        COMMENT_OFFSET = int(INCLUDE_COMMENTS)
        

        # Write each DataFrame to the appropriate sheet
        for sheetname, df in xls.items():
            
            # For the actual sheets where they will fill in their data, we dont want to write the column headers here
            df.to_excel(writer, sheet_name=sheetname, index=False, header = (not sheetname.startswith('tbl_')) )


        # Iterate over each sheet in the workbook
        for sheet in writer.sheets:
            # Access the active worksheet
            worksheet = writer.sheets[sheet]
            
            # Get the DataFrame corresponding to this sheet
            df = xls[sheet]
            
            if sheet in tabs_dict.keys():
                
                if INCLUDE_COMMENTS:
                    
                    # Grab the column comments from the database
                    column_comments = get_column_comments(sheet, eng)
                    comment_map = column_comments.set_index('column_name')['column_comment'].to_dict()

                    # Create a new DataFrame with comments as the first row
                    comment_row = [comment_map.get(col, '') for col in df.columns]
                    
                    worksheet.insert_rows(1)
                    
                    # Write the new row to the first row
                    for col_idx, value in enumerate(comment_row, start = 1):
                        cell = worksheet.cell(row = 1, column = col_idx)
                        cell.value = value
                        
                        # Apply text wrap to the comment cells
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        
                        # If app config has a custom column comment fill, set it
                        if CUSTOM_COLUMN_COMMENT_FILL is not None:
                            CUSTOM_COLUMN_COMMENT_FILL = str(CUSTOM_COLUMN_COMMENT_FILL).replace("#","").upper()
                            try:
                                cell = worksheet.cell(row = 1, column = col_idx)
                                cell.fill = PatternFill(start_color = CUSTOM_COLUMN_COMMENT_FILL, end_color = CUSTOM_COLUMN_COMMENT_FILL, fill_type = "solid")
                            except Exception as e:
                                print("Couldn't set custom column header fill - likely an error in app configuration")
                                print("Here is the exception message:")
                                print(e)
                        
                        # set font size for the comments
                        cell.font = Font(bold=False,size=COLUMN_COMMENT_FONT_SIZE)
                        
                        # Stick borders on the column comments
                        CUSTOM_COLUMN_COMMENT_BORDER_COLOR = str(CUSTOM_COLUMN_COMMENT_BORDER_COLOR).replace("#","").upper()
                        border_style = Side(border_style="thin", color=CUSTOM_COLUMN_COMMENT_BORDER_COLOR)
                        # Create the full border using the defined sides
                        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                        # stick the border on the cell
                        cell = worksheet.cell(row = 1, column = col_idx)
                        cell.border = border
                    
                # Write the column headers
                for col_idx, value in enumerate( list(df.columns), start = 1 ):
                    cell = worksheet.cell(row = 1 + COMMENT_OFFSET, column = col_idx)
                    cell.value = value
                    
                    max_length = len(str(value))
                    
                    # Give a little bit of a cushion - a bit more if there are comments
                    worksheet.column_dimensions[worksheet.cell(row= 1 + COMMENT_OFFSET, column = col_idx).column_letter].width = max_length + (5 * (1 + (COMMENT_OFFSET*2) ) )
                
                
                tmp_pkey_cols = tabs_dict.get(sheet).get('pkey_fields', [])
                tmp_constrained_cols = tabs_dict.get(sheet).get('constrained_columns', [])
                tmp_fkey_details = tabs_dict.get(sheet).get('foreign_key_detail', {})
                
                # Create a dictionary to map column names to their 1-based indices
                col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}
                
                # Get the column indices for primary key fields (1-based index)
                pkey_bold_col_indices = [col_indices[col] for col in tmp_pkey_cols]
                non_pkey_col_indices = [col_indices[col] for col in list(set(df.columns) - set(tmp_pkey_cols))]
                
                # Get the column indices for foreign key fields (1-based index)
                fkey_highlighted_cols = [col_indices[col] for col in tmp_constrained_cols]
                
                # Apply formatting for primary key columns (bold font)
                for col_idx in pkey_bold_col_indices:
                    worksheet.cell(row=1 + COMMENT_OFFSET, column=col_idx).font = PKEY_BOLD_FONT
                
                # Apply formatting for NON primary key columns (NON bold font)
                for col_idx in non_pkey_col_indices:
                    
                    worksheet.cell(row=1 + COMMENT_OFFSET, column=col_idx).font = Font(bold=False,size=COLUMN_HEADER_FONT_SIZE)
                    
                
                # Apply formatting for foreign key columns (gray highlight)
                for col_idx in fkey_highlighted_cols:
                    worksheet.cell(row=1 + COMMENT_OFFSET, column=col_idx).fill = FKEY_HIGHLIGHT
                    
                    # Get the relevant foreign key details
                    column_name = df.columns[col_idx - 1]
                    foreign_key_info = tmp_fkey_details.get(sheet, dict()).get(column_name)
                    
                    if foreign_key_info is not None:
                    
                        # col_idx is the index for the excel sheet, which is a 1 based index
                        referenced_table = tmp_fkey_details.get(sheet).get(df.columns[ col_idx - 1 ]).get('referenced_table')
                        referenced_column = tmp_fkey_details.get(sheet).get(df.columns[ col_idx - 1 ]).get('referenced_column')
                        
                        # Add a comment to the header indicating the lookup table
                        header_cell = worksheet.cell(row=1 + COMMENT_OFFSET, column=col_idx)
                        comment_text = f"References {referenced_table}.{referenced_column}"
                        header_cell.comment = Comment(text=comment_text, author="System")
                        
                        
                        
                        referenced_sheetname = quote_sheetname(referenced_table)
                        referenced_sheet_column_letter = get_column_letter(xls.get(referenced_table).columns.get_loc(referenced_column) + 1)
                        
                        # Find the last row in the referenced table's worksheet
                        referenced_sheet = writer.sheets[referenced_table]
                        max_ref_row = referenced_sheet.max_row
                        
                        dv = DataValidation(
                            type="list",
                            formula1=f"={referenced_sheetname}!${referenced_sheet_column_letter}$2:${referenced_sheet_column_letter}${max_ref_row}",
                            allow_blank = True
                        )
                        
                        
                        dv.error ='Your entry is not in the list'
                        dv.errorTitle = 'Invalid Entry'
                        
                        
                        # Convert column index to Excel column letter
                        col_letter = get_column_letter(col_idx)
                        
                        # Apply the validation to the entire column, starting from row 2 + COMMENT_OFFSET
                        dv.add(f"{col_letter}{2 + COMMENT_OFFSET}:{col_letter}1048576")
                        
                        worksheet.add_data_validation(dv)
                        
                        # Apply Conditional Formatting to highlight invalid entries
                        formula = f'=AND({col_letter}{2 + COMMENT_OFFSET}<>"", COUNTIF({referenced_sheetname}!${referenced_sheet_column_letter}$2:${referenced_sheet_column_letter}${max_ref_row},{col_letter}{2 + COMMENT_OFFSET})=0)'

                        worksheet.conditional_formatting.add(
                            f"{col_letter}{2 + COMMENT_OFFSET}:{col_letter}1048576",
                            FormulaRule(formula=[formula], fill=DATA_VALIDATION_ERROR_FILL)
                        )


                # Apply final styling for the column headers
                for col_idx in range(1, len(df.columns) + 1):  # Use 1-based indexing
                    
                    # define the cell object that we will apply styling to
                    cell = worksheet.cell(row=1 + COMMENT_OFFSET, column=col_idx)
                    
                    # Apply rotation and centering to all column headers
                    cell.alignment = COL_HEADER_ROTATION
                    
                    # If app config has a custom column header fill, set it - overriding the previously set ones for the foreign keys, etc.
                    if CUSTOM_COLUMN_HEADER_FILL is not None:
                        CUSTOM_COLUMN_HEADER_FILL = str(CUSTOM_COLUMN_HEADER_FILL).replace("#","").upper()
                        try:
                            cell = worksheet.cell(row = 1 + COMMENT_OFFSET, column = col_idx)
                            cell.fill = PatternFill(start_color = CUSTOM_COLUMN_HEADER_FILL, end_color = CUSTOM_COLUMN_HEADER_FILL, fill_type = "solid")
                        except Exception as e:
                            print("Couldn't set custom column header fill - likely an error in app configuration")
                            print("Here is the exception message:")
                            print(e)
                    
                        
                    
                    # Border
                    try:
                        # Apply custom border - default it to black
                        CUSTOM_COLUMN_HEADER_BORDER_COLOR = str(CUSTOM_COLUMN_HEADER_BORDER_COLOR).replace("#","").upper()
                        
                        border_style = Side(border_style="thin", color=CUSTOM_COLUMN_HEADER_BORDER_COLOR)
                        
                        # Create the full border using the defined sides
                        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

                        # stick the border on the cell
                        cell = worksheet.cell(row = 1 + COMMENT_OFFSET, column = col_idx)
                        cell.border = border
                        
                    except Exception as e:
                        print("Couldn't set column border - likely an error in app configuration")
                        print("Here is the exception message:")
                        print(e)
                            
            
            else:
                # Set the column widths based on max length in column
                for column_cells in worksheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[0:])
                    adjusted_width = max_length + 5  # Add cushion for a little extra space
                    worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

                # Apply filters to the specified header row
                if worksheet.max_row >= 0:  # Check if the header_row is within the data range
                    worksheet.auto_filter.ref = f"{worksheet.dimensions.split(':')[0]}:{worksheet.dimensions.split(':')[1]}"

    ############################################################################################################################
    ############################################################################################################################
    
    # set blob to the beginning
    excel_blob.seek(0)
    
    # Make a response object to set a custom cookie
    resp = make_response(send_file(excel_blob, as_attachment=True, download_name=f'{file_prefix}-TEMPLATE.xlsx'))

    # Set a cookie to let browser know that the file has been sent
    resp.set_cookie('template_file_sent', 'true', max_age=1)

    print("End Templater")

    return resp



